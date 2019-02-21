# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from textdistance import levenshtein


wb = load_workbook(filename='reestr.xlsx', read_only=True)
ws = wb.active
arr = [[cell.value for cell in row] for row in ws.rows]
i = 2
cur = 1

'''for i in range(10, 25):
    print(arr[i][7:10], arr[i][13], arr[i][15:19])
exit()'''
D = dict()
D2 = dict()
C = open('collizions.txt', 'r')
C = C.readline().split()
used = []
 
while i < len(arr):
    if arr[i][0] is None:
        i += 1
        continue
    if levenshtein.distance(str(arr[i][0]), str(cur)) > 100 or arr[i][1] is None or len(str(arr[i][2])) == 1:
        i += 1
        continue
    if levenshtein.distance(str(arr[i][12]), 'Автобус') > 5:
        i += 1
        continue
    name = str(arr[i][1])
    name = name.replace('/', '')
    name = name.replace('в', '8')
    name = name.replace('s', '5')
    name = name.replace('з', '3')
    name = name.replace('о', '0')
    name = name.replace('ы', '61')
    name = name.upper()
    
    if name in D2:
        D2[name] += 1
    else:
        D2[name] = 1
    
    if name in C:
        '''tt = arr[i][2].split('-')[0]
        tt = tt.replace('.', '')
        tt = tt.replace(',', '')
        tt = tt.replace('"', '')
        tt = tt.replace('\n', ' ')
        name += ' ' + tt[:len(tt) - 1]'''
        if 'Зеленоград' in arr[i][5]:
            name = 'З-' + name
        elif name in D:
            name = 'НМ-' + name
        #else:
        #    name += ' ' + str(i)
    #print(arr[i])
    #break
    
    if name in D:
        D[name] += 1
    else:
        D[name] = 1
    
    used.append(name)
    fout = open('reestr/' + name + '.txt', 'w')
    t1 = arr[i][7:10]
    t3 = str(arr[i][13])
    #print(arr[i][1], name, arr[i][13])
    #if '1' in name.split():
    #    print(arr[i][0], name, arr[i][13])
    for j in range(len(t1)):
        t1[j] = str(t1[j])
        if t1[j] == '40':
            t1[j] = '4,0'
        t1[j] = t1[j].replace('в', '8')
        t1[j] = t1[j].replace('s', '5')
        t1[j] = t1[j].replace('з', '3')
        t1[j] = t1[j].replace('о', '0')
        t1[j] = t1[j].replace('а', '4')
        t1[j] = t1[j].replace('ы', '61')
    for j in range(len(t1)):
        if '\n' in t1[j]:
            t1[j] = t1[j].replace('\n', '')
    t2 = arr[i][15:19]
    for j in range(len(t2)):
        if t2[j] is None:
            t2[j] = 0
        t2[j] = str(t2[j])
        t2[j] = t2[j].replace('в', '8')
        t2[j] = t2[j].replace('s', '5')
        t2[j] = t2[j].replace('з', '3')
        t2[j] = t2[j].replace('о', '0')
        t2[j] = t2[j].replace('а', '4')
        t2[j] = t2[j].replace('ы', '61')
    t3 = t3.replace('о', '0')
    t3 = t3.replace('s', '5')
    t3 = t3.replace('з', '3')
    t3 = t3.replace('в', '8')
    t3 = t3.replace('а', '4')
    t3 = t3.replace('ы', '61')
    i += 1
    cur += 1
    print(' '.join(t1) + ';' + t3 + ';' + ' '.join(list(map(str, t2))), file=fout)
    fout.close()


wb = load_workbook(filename='reestr2.xlsx', read_only=True)
ws = wb.active
arr = [[cell.value for cell in row] for row in ws.rows]
i = 2
cur = 1
while i < len(arr):
    if arr[i][0] is None or arr[i][1] is None:
        i += 1
        continue
    if levenshtein.distance(str(arr[i][0]), str(cur)) > 100:
        i += 1
        continue
    if levenshtein.distance(str(arr[i][13]), 'Автобус') > 2:
        i += 1
        continue
    name = str(arr[i][2])
    name = name.upper()
    
    if name in D2:
        D2[name] += 1
    else:
        D2[name] = 1
    
    if name in C:
        if 'Зеленоград' in arr[i][5]:
            name = 'З-' + name
        elif name in D:
            name = 'НМ-' + name
    if name in used:
        name += '-2'
    
    if name in D:
        D[name] += 1
    else:
        D[name] = 1
    
    used.append(name)
    fout = open('reestr/' + name + '.txt', 'w')
    t1 = arr[i][8:11]
    t3 = str(arr[i][14])
    for j in range(len(t1)):
        t1[j] = str(t1[j])
    for j in range(len(t1)):
        if '\n' in t1[j]:
            t1[j] = t1[j].replace('\n', '')
    t2 = arr[i][16:20]
    for j in range(len(t2)):
        if t2[j] is None:
            t2[j] = 0
    i += 1
    cur += 1
    print(' '.join(t1) + ';' + t3 + ';' + ' '.join(list(map(str, t2))), file=fout)
    fout.close()


'''print(D2)
coll = open('collizions.txt', 'w')
for i in D2:
    if D2[i] > 1:
        print(i, end=' ', file=coll)
        print(i, D2[i])
coll.close()'''